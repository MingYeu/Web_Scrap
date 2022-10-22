#Search Product
from fileinput import filename
from lib2to3.pgen2.token import EQUAL
from xml.etree.ElementTree import Comment
from bs4 import BeautifulSoup
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.webdriver.common.by import By
import re
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path
from datetime import date
import numpy

#Pie Chart
from matplotlib import pyplot as plt
import numpy as np

#UI
import PySimpleGUI as sg

time_getToUrl = 10
sheet_name = ''

new_file = False
new_sheet = False

is_shopee = True
is_lazada = False

scrolling_time_window = 20
time_scrollWindow = 1
time_load_individual_page = 5

num_web_scrap_glob = 60
start_web_scrap_glob = 0

# Date
today = date.today()
date = today.strftime("%d/%m/%Y")

# Excel
excelSpace = ['b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'y', 'z', 
              'aa', 'ab', 'ac', 'ad', 'ae', 'af', 'ag', 'ah', 'ai', 'aj', 'ak', 'al', 'am', 'an', 'ao', 'ap', 'aq', 'ar', 'as', 'at', 'au', 'av', 'aw', 'ay', 'az']


webdriver = webdriver.Chrome(executable_path=ChromeDriverManager().install())

def search_all_product(file_name, keyword):
    # ################################Excel#########################################

    # if check_file_close()==False:
    #     return 
    print(file_name)
    print(keyword)
    workbook = load_workbook('database/' + file_name)
    print('yes')
    worksheet = workbook[keyword]
    print('ok')

    # if os.path.exists(file_name):
    #     workbook = load_workbook(file_name)
    #     exists = False
    #     check_worksheet = workbook.sheetnames
    #     for check_sheet in check_worksheet:
    #         if check_sheet == keyword:
    #             exists = True
                
    #     if exists:
    #         worksheet = workbook[keyword]
    #     else:
    #         new_worksheet = workbook.create_sheet(keyword)
    #         worksheet = workbook.active = workbook[keyword]
    #         new_sheet = True

    # else:
    #     workbook = Workbook()
    #     worksheet = workbook.active
    #     # worksheet = workbook["Skateboard"]
    #     check_worksheet = workbook.sheetnames
    #     worksheet.title = keyword


    # #################################Web Scraping#########################################

    # webdriver = webdriver.Chrome(executable_path=ChromeDriverManager().install())
    
    # print(new_sheet)
    Website_no_error = False
    # if new_sheet:
    # url = ''
    # if is_shopee:
    #     url = 'https://shopee.com.my/search?keyword=' + keyword
    #     second_url = 'https://shopee.com.my/search?keyword=' + keyword + '&page=1'
    #     Website_no_error = True
    # elif is_lazada:    
    #     url = 'https://www.lazada.com.my/catalog/?spm=a2o4k.home.search.1.75f82e7ex5Ek0y&q=Skateboard&_keyori=ss&from=search_history&sugg=Skateboard_0_1'
    #     Website_no_error = True
    # else:
    #     return "Shopee Or Lazada Error"


    url = 'https://shopee.com.my/search?keyword=' + keyword
    second_url = 'https://shopee.com.my/search?keyword=' + keyword + '&page=1'
        
    # if Website_no_error == True:
    webdriver.get(url)
    
    time.sleep(time_getToUrl)

    for i in range(scrolling_time_window): 
        webdriver.execute_script("window.scrollBy(0, 250)")
        time.sleep(time_scrollWindow)


    html = BeautifulSoup(webdriver.page_source,'html.parser')

    # all_product = ''
    # if is_shopee:
    #     all_product = 'col-xs-2-4 shopee-search-item-result__item'
    # elif is_lazada: 
    #     all_product = '_95X4G'

    all_product = 'col-xs-2-4 shopee-search-item-result__item'
    # ################# All Products ########################
    worksheet.append([])

    num_web_scrap = num_web_scrap_glob
    start_web_scrap = start_web_scrap_glob
    
    # print(str(start_web_scrap) + ">" + str(num_web_scrap))
    productDivs = html.findAll('div', attrs={'class' : all_product})
    for div in productDivs:
        # print(str(start_web_scrap) + ">" + str(num_web_scrap))
        if start_web_scrap >= num_web_scrap:
            break
        # print(div.a['href'])
        web_scrap_prod = div.a['href']
        start_web_scrap += 1
        worksheet.append([web_scrap_prod])

        
    workbook.save('database/' + file_name)
    search_each_product(file_name)

def search_each_product(file_name):
    ################# Individual Product ########################
    path = 'database/' + file_name
    workbook = load_workbook(path)
    sheet_name = workbook.sheetnames
    
    for keyword in sheet_name:    
        worksheet = workbook[keyword]

        excel_space = excelSpace
        last_space = 'a'
        excel_row = '2'

        check_date = 'a1'
        # Check last space
        for i in excel_space:
            column_excel = i + '1'
            last_space = i
            if worksheet[column_excel].value == None:
                check_date = excel_space.index(i)
                break
        
        check_date = excel_space[check_date - 1] + '1'
        # print(excel_space[check_date - 1])
        
        test = worksheet[check_date].value
        testdate = test.strftime("%d/%m/%Y")
        # print(date)
        
        if testdate != date:
            # Insert Date
            worksheet[last_space + '1'] = today

            # Class Name
            name_class = '_2rQP1z'
            comment_class = '_3y5XOB'
            sold_class = 'HmRxgn'
            price_class = '_2Shl1j'

            # name_class = ''
            # comment_class = ''
            # sold_class = ''
            # price_class = ''

            # Load Product
            span_sub1 = "<span>"
            span_sub2 = "</span>"
            div_class_sub1 = '">'
            div_class_sub2 = '</'


            for row in worksheet.iter_rows(values_only=True):
                product_link  = row[0]

                if product_link == None:
                    continue

                webdriver.get('https://shopee.com.my' + product_link)
                
                time.sleep(time_load_individual_page)
                
                html = BeautifulSoup(webdriver.page_source,'html.parser')

                # Name
                # product_name = ''
                # uls = html.findAll('div', attrs = {'class':name_class})
                # for div in uls:
                #     product_name_span = str(div.span)
                #     # print(product_name_span)
                #     for idx in range(product_name_span.index(span_sub1) + len(span_sub1) , product_name_span.index(span_sub2)):
                #         product_name = product_name + product_name_span[idx]
                #     # print(product_name)
                
                # star & comments 
                star = ''
                comments = ''
                result = ''
                loop_time = 0
                # uls = html.findAll('div', attrs = {'class':comment_class})
                # if uls == []:
                #     star = str(0)
                #     comments = str(0)
                # else:
                #     starcomment = str(uls).split(",")
                    
                #     for stco in starcomment:
                #         for idx in range(stco.index(div_class_sub1) + len(div_class_sub1) , stco.index(div_class_sub2)):
                #             result = result + stco[idx]
                #         if loop_time == 0:
                #             star = result
                #         else:
                #             comments = result
                #         loop_time += 1
                #     result = ''
                
                # sold
                sold = ''
                uls = []
                uls = str(html.findAll('div', attrs = {'class':sold_class}))
                # if len(uls) != 0:
                # if 'div' in uls:
                #     print('exists')
                # print(uls)
                # print(len(uls))
                if 'div' in uls:    
                    for idx in range(uls.index(div_class_sub1) + len(div_class_sub1) , uls.index(div_class_sub2)):
                        result = result + uls[idx]
                    sold = result
                    result = ''
                    
                    # price
                    price = ''
                    uls = str(html.findAll('div', attrs = {'class':price_class}))
                    for idx in range(uls.index(div_class_sub1) + len(div_class_sub1) , uls.index(div_class_sub2)):
                        result = result + uls[idx]
                    price = result
                    result = ''
                    
                    # result_store = product_name + '&&&&' + star + '&&&&' + comments + '&&&&' + sold + '&&&&' + price
                    result_store = sold + '&&&&' + price
                    # print(result_store)
                    
                    # Place to Store
                    worksheet[str(last_space + excel_row)] = result_store
                    excel_row = str(int(excel_row) + 1)

                    product_name = ''
                else:
                    sold = '0' 
                    price = '0'
                    result_store = sold + '&&&&' + price  
                    worksheet[str(last_space + excel_row)] = result_store
                    excel_row = str(int(excel_row) + 1)  
                
            workbook.save(path)

def create_pie_chart(file_name):

    # Retrieve from Excel
    workbook = load_workbook(file_name)
    sheet_name = workbook.sheetnames
    count_sheet = len(sheet_name)
    # print(count_sheet)
    # print(sheet_name)
    
    total_price_sold = 0
    sold_amount = 0
    sold_amount_price = 0
    total_sold_amount = []
    data_value = []
    for x in sheet_name:
        worksheet = workbook[x]
        
        last = latest_column(worksheet)
        # print(last)
        
        if last == 'a':
            return 'There is no value inside'
        
        sold_amount_price = 0
        one_price_sold = 0
        
        for y in range(2, 100):
            excel_location = last + str(y)
            # print(excel_location)
            
            price_retrieve = worksheet[excel_location].value
           
            if price_retrieve == None:
                break
            arrya = price_retrieve.split("&&&&")
            one_price_sold = arrya[-1]
            sold_amount = arrya[-2]
            
            # Price
            # one_price_sold = one_price_sold.replace(' ', '')
            # one_price_sold = one_price_sold.replace('RM', '')
            # one_price_sold = one_price_sold.replace(',', '')
            
            # if "-" in one_price_sold:
            #     got_dash = one_price_sold.split("-")  
            #     one_price_sold = got_dash[1] 
            
            # Amount
            one_sold = 0
            if "k" in sold_amount:
                got_k = sold_amount.replace('k', '')
                # print(got_k)
                one_sold = float(got_k) * 1000
                # print(one_sold)
                # print(type(one_sold))
                # sold_amount_price = sold_amount_price + one_sold
            else:
                one_sold = float(sold_amount) 
            
            sold_amount_price = sold_amount_price + one_sold
            
            # def isfloat(num):
            #     try:
            #         float(num)
            #         return True
            #     except ValueError:
            #         return False
            
            # Price
            # if isfloat(one_price_sold) == True:
            #     total_price_sold = total_price_sold + float(one_price_sold)
            #     print(type(total_price_sold))
            #     print(total_price_sold)
        
        total_sold_amount.append(sold_amount_price)
        # data_value.append(total_price_sold)  
        total_price_sold = 0
        sold_amount_price = 0
        # print(total_sold_amount)
        # print(data_value)
        
    # Create Pie Chart
    myLabel = sheet_name
    
    # myexplode = []

    # for z in range(0, count_sheet):
    #     if z == 0:
    #         myexplode = [0]
    #     elif z == 1:
    #         myexplode = [0.2, 0]
    #     elif z == 2:
    #         myexplode = [0.3, 0.2, 0]
    #     else:
    #         myexplode = [0] * 5
    #         myexplode[0] = 0.4
    #         myexplode[1] = 0.3
    #         myexplode[2] = 0.2
        
    # print(myLabel)
    # print(data_value)
    # print(myexplode)
    
       
    
    og_array = np.array(total_sold_amount)
    total_sold_amount = og_array.astype(int)
    
    # Ascending Order    
    # a = np.array(myLabel)
    # b = np.array(total_sold_amount)
    
    # new_order = np.lexsort([b, a])
    # myLabel = a[new_order]
    # total_sold_amount = b[new_order]

    
    
    # sizes = numpy.array(data_value)
    # def absolute_value(val):
    #     a  = numpy.round(val/100.*sizes.sum(), 0)
    #     return a
    # total_pie = sum(data_value)
    
    # Creating plot
    fig = plt.figure(figsize =(10, 7))
    # plt.pie(data_value, labels = myLabel, explode = myexplode, autopct='%1.1f%%', startangle=90)
    
    plt.pie(total_sold_amount, labels = myLabel, autopct='%1.1f%%', startangle=90)
    plt.legend(title = file_name)
    
    cellText = []
    for i in range(len(sheet_name)):
        # item_display = [sheet_name[i], str(total_sold_amount[i]), str(data_value[i])]
        item_display = [sheet_name[i], str(total_sold_amount[i])]
        cellText.append(item_display)
    # print(cellText)
    #Table
    data_table = plt.table(
        cellText=cellText,
        # rowLabels=['Jan', 'Feb',],
        # colLabels=['Product', 'Sold Amount (Quantity)', 'Total Sales (RM)'],
        colLabels=['Product', 'Sold Amount (Quantity)'],
        # rowColours=colors,
        colWidths=[0.35, 0.35, 0.35],
        loc='bottom'
    )
    
    data_table.scale(1.5, 3)
    data_table.set_fontsize(12)
    plt.xticks(())
    plt.subplots_adjust(left=0.1, bottom=0.25)

    # show plot
    plt.show()


def create_fast_pie_chart_date_select(file_name, keyword):
    
    workbook = load_workbook(file_name)
    worksheet = workbook[keyword]
    
    all_date = []
    excel_space = excelSpace
    for i in excel_space:
        column_excel = i + '1'
        last_space = i
        if worksheet[column_excel].value == None:
            check_date = excel_space.index(i)
            break
        
        date = worksheet[column_excel].value.strftime("%d/%m/%Y")
        all_date.append(date)
    
    layout1 = [[sg.Text('Initial Date', size=(10,1)),sg.Combo(all_date,key='initial_date')],
               [sg.Text('Start Date', size=(10,1)),sg.Combo(all_date,key='start_date')],
               [sg.Text('End Date', size=(10,1)),sg.Combo(all_date,key='end_date')],
           [sg.Button('Start Analysis')]]

    window = sg.Window("Select Date", layout1, modal=True)
    choice = None
    while True:
        event, values = window.read()
        if event == "Exit" or event == sg.WIN_CLOSED:
            break
        elif event == 'Start Analysis':
            
            initial_po = all_date.index(values['initial_date'])
            initial_date = excel_space[initial_po]
            # print(worksheet[initial_date].value)
            
            start_po = all_date.index(values['start_date'])
            start_date = excel_space[start_po] 
            
            end_po = all_date.index(values['end_date'])
            end_date = excel_space[end_po]
            
            create_fast_pie_chart(file_name, keyword, initial_date, start_date, end_date)
            
    window.close()

def create_fast_pie_chart(file_name, keyword, initial_date, start_date, end_date):
    
    workbook = load_workbook(file_name)
    worksheet = workbook[keyword]
    
    all_date_app = []
    all_date_app.append(initial_date)
    all_date_app.append(start_date)
    all_date_app.append(end_date)
    
    sold_amount = 0
    sold_amount_price = 0
    total_sold_amount = []
    
    
    for x in all_date_app:
        one_price_sold = 0
        sold_amount_price = 0
        
        excel_col_loc = x
        
        
        for y in range(2, 100):
            excel_location = x + str(y)
            
            price_retrieve = worksheet[excel_location].value
            
            if price_retrieve == None:
                break
            arrya = price_retrieve.split("&&&&")
            one_price_sold = arrya[-1]
            sold_amount = arrya[-2]
            
            # Amount
            one_sold = 0
            if "k" in sold_amount:
                got_k = sold_amount.replace('k', '')

                one_sold = float(got_k) * 1000

            else:
                one_sold = float(sold_amount) 
            
            sold_amount_price = sold_amount_price + one_sold
            
            
        total_sold_amount.append(int(sold_amount_price)) 
        total_price_sold = 0
        sold_amount_price = 0
        
    print(total_sold_amount)
    
    
    # Start Analysis
    upOrDown = []
    length = len(total_sold_amount)
    
    for i in reversed(total_sold_amount):
        if i == total_sold_amount[0]:
            break
        upOrDown.append(i - total_sold_amount[length - 2])
        length = length - 1
    
    percen_up = 0
    length = len(upOrDown)
    
    for i in upOrDown:
        if i == upOrDown[-1]:
            break
        result = (i - upOrDown[length - 1]) / upOrDown[length - 1] * 100
        # percen_up.append(result)
        percen_up = result
        
    start = worksheet[start_date + '1'].value.strftime("%d/%m/%Y") 
    end = worksheet[end_date + '1'].value.strftime("%d/%m/%Y") 
        
    sg.popup('Result',      
            str(start) + ' to ' + str(end) + ' has increase ' + str(percen_up) + '%' )

def latest_column(worksheet):
    excel_space = excelSpace
    count = 0

    # Check last space
    for i in excel_space:
        column_excel = i + '1'
        count += 1
        if worksheet[column_excel].value == None:
            break   
    return excel_space[count - 2]

# def check_file_close():
#     f = open(file_name)
#     if f.closed:
#         # print('close')
#         return True
#     else:
#         # print('open')
#         return False


# keyword = 'Skateboards'
# file_name = 'Infor.xlsx'

# search_all_product(file_name, keyword)
# search_each_product(file_name)
# create_pie_chart(file_name)

Category = os.listdir(os.getcwd() + '\database')
Item = []


def analysis():
    layout1 = [[sg.Text('Category', size=(10,1)),sg.Combo(Category,key='board_report')],
           [sg.Button('Start Analysis')]]
    
    layout5 = [[sg.Text('Category', size=(10,1)),sg.Combo(Category,key='board_fast_report'), sg.Button('Check')],
           [sg.Text('Item', size=(10,1)),sg.Input('',key='board_fast_item')],
           [sg.Button('Start Fast Analysis')]]

    #Define Layout with Tabs         
    analysisgrp = [[sg.TabGroup([[sg.Tab('Analysis Report', layout1, title_color='Black',border_width =10, background_color='white', element_justification= 'center'),
                            sg.Tab('Analysis Fast Report', layout5, title_color='Black',border_width =10, background_color='white', element_justification= 'center'),
                            ]], 
                        tab_location='centertop',
                        title_color='Black', tab_background_color='White',selected_title_color='Gray',
                        selected_background_color='White', border_width=5)]]  
              
    window =sg.Window("Business Analysis",analysisgrp)
    
    while True:          
        event, values = window.Read()
        if event in (None, 'Exit'):
            break
        if event == 'Start Analysis':
            create_pie_chart('database/' + values['board_report'])
            
        elif event == 'Check':
            workbook = load_workbook('database/' + str(values['board_fast_report']))
            Item = workbook.sheetnames
            sg.popup('Exist Item: ',      
                Item )
            
        elif event == 'Start Fast Analysis':
            create_fast_pie_chart_date_select('database/' + str(values['board_fast_report']), values['board_fast_item'])

    window.close()  
    

def collect():
    layout2=[[sg.Text('Category', size=(15,1)),sg.Combo(Category,key='board_collect')], 
         [sg.Button('Start Collect')]]
    
    collectgrp = [[sg.TabGroup([[sg.Tab('Collect Key', layout2,title_color='Black',background_color='white', element_justification= 'center'),
                         ]], 
                       tab_location='centertop',
                       title_color='Black', tab_background_color='White',selected_title_color='Gray',
                       selected_background_color='White', border_width=5)]]         

    window =sg.Window("Collect Data",collectgrp)

    
    # Event Loop
    while True:          
        event, values = window.Read()
        if event in (None, 'Exit'):
            break
        if event == 'Start Collect':
            search_each_product(values['board_collect'])

    window.close()    

def add_new_item():
    layout3= [[sg.Text('Existing Category', size=(15,1)),sg.Combo(Category,key='board_add_exist')],
            [sg.Text('Keyword', size=(15,1)),sg.Input('',key='exist_cat_new_key')],
            [sg.Button('Add Existing Category')]]

    layout4= [[sg.Text('Add Category', size=(10,1)),sg.Input('',key='eAddCategory')],
            [sg.Text('Keyword', size=(15,1)),sg.Input('',key='new_key')],
            [sg.Button('Add Category')]]


    #Define Layout with Tabs         
    addgrp = [[sg.TabGroup([[sg.Tab('Add Existing Category', layout3,title_color='Black',background_color='white', element_justification= 'center'),
                            sg.Tab('Add Category', layout4,title_color='Black',background_color='white', element_justification= 'center'),]], 
                        tab_location='centertop',
                        title_color='Black', tab_background_color='White',selected_title_color='Gray',
                        selected_background_color='White', border_width=5)]]  
            
    #Define Window
    window =sg.Window("Add New Group",addgrp)

    # Event Loop
    while True:          
        event, values = window.Read()
        if event in (None, 'Exit'):
            break  
        if event == 'Add Existing Category':
            check_file_exist = os.listdir(os.getcwd() + '\database')  
            workbook = load_workbook('database/' + str(values['board_add_exist']))
            sheet_name = workbook.sheetnames
            names_lower = [name.lower() for name in sheet_name]
            
            if (values['exist_cat_new_key']).lower() in names_lower:
                sg.popup('Error',      
                'The key already exist' )
            else:
                workbook.create_sheet(values['exist_cat_new_key'])
                workbook.save('database/' + values['board_add_exist'])
                search_all_product(str(values['board_add_exist']), str(values['exist_cat_new_key']))
                
        elif event == 'Add Category':
            check_file_exist = os.listdir(os.getcwd() + '\database')
            if (values['exist_cat_new_key']).lower() in check_file_exist:
                sg.popup('Error',      
                'Category already exist' )
            else:
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = (values['new_key']).lower()
                workbook.save('database/' + str((values['eAddCategory']).lower()) + '.xlsx')
                search_all_product(str(values['eAddCategory']), str(values['new_key']))
                
    #access all the values and if selected add them to a string
    window.close()    


# search_all_product(str(values['eAddCategory']), str(values['new_key']))
# search_all_product('toy', 'skateboards')


# Main
layout=[[sg.Button('Analysis Report'),sg.Button('Collect Data')], 
         [sg.Button('Add New Item')]]
    
maingrp = [[sg.TabGroup([[sg.Tab('Business Analysis', layout,title_color='Black',background_color='white', element_justification= 'center'),
                        ]], 
                    tab_location='centertop',
                    title_color='Black', tab_background_color='White',selected_title_color='Gray',
                    selected_background_color='White', border_width=5)]]         

window =sg.Window("Business Analysis",maingrp)


# Event Loop
while True:          
    event, values = window.Read()
    if event in (None, 'Exit'):
        break
    if event == 'Analysis Report':
        analysis()
    elif event == 'Collect Data':
        collect()
    elif event == 'Add New Item':
        add_new_item()        
        
window.close()    

